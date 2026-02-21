"""
Excel/CSV Loader — row-per-document strategy.

Problem with naive Excel loading: flatten 500 rows into one Document.
The LLM sees an undifferentiated wall of numbers.

Solution: each row becomes its own Document. Column names become context
for every row: "Product: A99 | Region: APAC | Price: $120 | Stock: 45"

This means:
- Semantic search for "product A99 pricing" finds the exact row
- Metadata extraction pulls product_id directly from the serialized row
- The LLM can cite specific rows in its answer

Multiple sheets: each sheet is processed independently.
Header detection: assumes row 0 is headers (override via skip_rows param).
Empty rows: skipped.
"""
from pathlib import Path

from langchain_core.documents import Document

from app.loaders.base import BaseLoader


class ExcelLoader(BaseLoader):
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls", ".csv"]

    def load(self, file_path: Path) -> list[Document]:
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self._load_csv(file_path)
        return self._load_excel(file_path)

    def _load_excel(self, path: Path) -> list[Document]:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        docs = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h or f"col_{i}").strip() for i, h in enumerate(rows[0])]
            for row_idx, row in enumerate(rows[1:], start=2):
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                parts = [
                    f"{headers[i]}: {str(v).strip()}"
                    for i, v in enumerate(row)
                    if v is not None and str(v).strip()
                ]
                if not parts:
                    continue
                docs.append(Document(
                    page_content=" | ".join(parts),
                    metadata={
                        "source": path.name,
                        "sheet": sheet_name,
                        "row": row_idx,
                        "is_table": True,
                    },
                ))
        return docs

    def _load_csv(self, path: Path) -> list[Document]:
        import csv
        docs = []
        with open(path, encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f)
            for row_idx, row in enumerate(reader, start=2):
                parts = [f"{k}: {v.strip()}" for k, v in row.items() if v and v.strip()]
                if not parts:
                    continue
                docs.append(Document(
                    page_content=" | ".join(parts),
                    metadata={"source": path.name, "row": row_idx, "is_table": True},
                ))
        return docs
